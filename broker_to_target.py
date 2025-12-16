# -*- coding: utf-8 -*-
"""
broker_to_target.py — точка входа CLI.

Требует:
  - helpers.py
  - core.py

Примеры:
  python broker_to_target.py --input report.xlsx --output target.xlsx
  python broker_to_target.py --input report.xlsx --output target.xlsx \
      --include-cash --alloc-commission --sheet "Отчет" --debug
"""

from __future__ import annotations

import argparse
from typing import List

import openpyxl
import pandas as pd

from core import (
    TargetRow,
    parse_portfolio_isin_map,
    parse_trades_to_target,
    parse_cash_to_target,
)

# ---------------------------- Утилиты сохранения ------------------------------

def rows_to_dataframe(rows: List[TargetRow]) -> pd.DataFrame:
    """
    Преобразует TargetRow -> DataFrame с колонками в требуемом порядке.
    """
    columns = [
        "Event", "Date", "Symbol", "Price", "Quantity", "Currency",
        "FeeTax", "Exchange", "NKD", "FeeCurrency", "DoNotAdjustCash", "Note"
    ]
    data = [{col: getattr(r, col) for col in columns} for r in rows]
    return pd.DataFrame(data, columns=columns)


def save_xlsx(df: pd.DataFrame, path: str, sheet_name: str = "Sheet1") -> None:
    """
    Сохранение в XLSX без индексов. По умолчанию engine='openpyxl'.
    """
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

# ---------------------------- Main -------------------------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Конвертация брокерского XLSX в целевой формат target.xlsx"
    )
    p.add_argument("--input", required=True, help="Путь к исходному XLSX отчёту")
    p.add_argument("--output", required=True, help="Путь к целевому XLSX (например, target.xlsx)")
    p.add_argument("--sheet", default=None, help="Имя листа в отчёте (по умолчанию — первый)")
    p.add_argument("--include-cash", action="store_true", help="Добавлять данные из «Движение ДС»")
    p.add_argument("--alloc-commission", action="store_true", help="Распределять комиссию «итого по выпуску» пропорционально суммам сделок")
    p.add_argument("--locale-comma", action="store_true", default=True, help="Числа в отчёте с запятой (по умолчанию включено)")
    p.add_argument("--no-locale-comma", dest="locale_comma", action="store_false", help="Отключить замену запятых на точки при парсинге чисел")
    p.add_argument("--sort", action="store_true", help="Сортировать результат по дате (Date) и Event")
    p.add_argument("--debug", action="store_true", help="Печать отладочной информации")
    return p


def main():
    args = build_arg_parser().parse_args()

    # Читаем книгу
    wb = openpyxl.load_workbook(args.input, data_only=True, read_only=True)
    try:
        ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    except KeyError:
        wb.close()
        raise SystemExit(f"Лист '{args.sheet}' не найден. Доступные листы: {wb.sheetnames}")

    if args.debug:
        print(f"Открыт файл: {args.input}, лист: {ws.title}")

    # 1) Карту ISIN из портфеля собираем один раз
    name_to_isin, regnum_to_isin = parse_portfolio_isin_map(ws, debug=args.debug)

    # 2) Сделки
    trade_rows = parse_trades_to_target(
        ws,
        name_to_isin=name_to_isin,
        alloc_commission=args.alloc_commission,
        locale_comma=args.locale_comma,
        debug=args.debug,
    )
    if args.debug:
        print(f"Сделки: строк -> {len(trade_rows)}")

    # 3) Движение ДС (по необходимости)
    cash_rows: List[TargetRow] = []
    if args.include_cash:
        cash_rows = parse_cash_to_target(
            ws,
            name_to_isin=name_to_isin,
            regnum_to_isin=regnum_to_isin,
            locale_comma=args.locale_comma,
            debug=args.debug,
        )
        if args.debug:
            print(f"ДДС: строк -> {len(cash_rows)}")

    wb.close()

    # 4) Объединяем
    all_rows: List[TargetRow] = trade_rows + cash_rows

    # 5) Сортировка (по желанию)
    df = rows_to_dataframe(all_rows)
    if args.sort and not df.empty:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df.sort_values(by=["Date", "Event"], inplace=True, kind="mergesort")  # стабильная сортировка
        df["Date"] = df["Date"].dt.strftime("%Y-%m-%d")

    # 6) Сохранение
    save_xlsx(df, args.output)
    if args.debug:
        print(f"Готово. Записано {len(df)} строк в {args.output}")


if __name__ == "__main__":
    main()