# -*- coding: utf-8 -*-
"""
core.py — основные функции преобразования отчёта:
- парсинг портфеля (ISIN карта),
- парсинг сделок,
- парсинг движения денежных средств.

Зависит от helpers.py.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from helpers import (
    norm_text,
    parse_decimal,
    parse_date,
    normalize_currency,
    match_header_indices,        # не обязательно, но пригодится для доп. сценариев
    find_section_title,
    find_header_below,
    to_exchange,
    norm_key,
    isin_from_text,
)

# ---------------------------- Константы/регэкспы -----------------------------

SECTION_TRADES_RE = re.compile(r"сделки\s+с\s+ценн(ыми|ыми)\s+бумагами", re.IGNORECASE)
SECTION_PORTFOLIO_RE = re.compile(r"состояние\s+портфеля\s+ценн(ых|ыми)\s+бумаг", re.IGNORECASE)
SECTION_CASH_RE = re.compile(r"движение\s+денежных\s+средств.*за\s+отчетный\s+период", re.IGNORECASE)
SECTION_CLEARANCE_RE = re.compile(r"исполнение\s+обязательств\s+по\s+сделке", re.IGNORECASE)

TRADE_TYPES_MAP: Dict[str, str] = {
    "покупка": "Buy",
    "продажа": "Sell",
}

HEADER_KEYS_TRADES: Dict[str, List[str]] = {
    "deal_no": ["номер сделки"],
    "deal_date": ["дата сделки"],
    "deal_time": ["время сделки"],
    "deal_type": ["вид сделки"],
    "price": ["цена одной цб", "цена одной", "цена цб"],
    "price_ccy": ["валюта цены"],
    "qty": ["количество цб", "количество", "кол-во цб"],
    "nkd": ["нкд"],
    "amount": ["сумма сделки", "сумма"],
    "amount_ccy": ["валюта суммы"],
    "comm_ts": ["комиссия тс", "комиссия торговой системы"],
    "comm_broker": ["комиссия брокера"],
    "place": ["место совершения сделки", "место сделки", "площадка"],
}

HEADER_KEYS_PORTFOLIO: Dict[str, List[str]] = {
    "name": ["наименование цб", "наименование", "инструмент"],
    "isin": ["isin"],
}

HEADER_KEYS_CASH: Dict[str, List[str]] = {
    "op_no": ["№ операции", "номер операции", "№ операции"],
    "date": ["дата"],
    "type": ["тип операции"],
    "amount": ["сумма"],
    "ccy": ["валюта"],
    "comment": ["комментарий", "примечание"],
}

# ---------------------------- Модель строки назначения ------------------------

@dataclass
class TargetRow:
    Event: str
    Date: str
    Symbol: Optional[str]
    Price: Optional[float]
    Quantity: Optional[float]
    Currency: Optional[str]
    FeeTax: Optional[float]
    Exchange: Optional[str]
    NKD: Optional[float]
    FeeCurrency: Optional[str]
    DoNotAdjustCash: Optional[str]
    Note: Optional[str]

# ---------------------------- Портфель / ISIN --------------------------------

def parse_portfolio_isin_map(ws: Worksheet, debug: bool = False) -> Dict[str, str]:
    """
    Из раздела «СОСТОЯНИЕ ПОРТФЕЛЯ ЦЕННЫХ БУМАГ» строит карту:
      { norm_key(Наименование ЦБ) -> ISIN }
    """
    title_row = find_section_title(ws, SECTION_PORTFOLIO_RE)
    if title_row is None:
        if debug:
            print("Портфель: раздел не найден.")
        return {}

    header_row, idx = find_header_below(ws, title_row, HEADER_KEYS_PORTFOLIO, needed=["name", "isin"])
    name_col = idx["name"] + 1
    isin_col = idx["isin"] + 1

    mapping: Dict[str, str] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(r, name_col).value
        isin = ws.cell(r, isin_col).value
        if not (name or isin):
            # Возможный конец таблицы — пустая строка
            row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if all(v in (None, "", " ") for v in row_vals):
                break
            continue
        name_s = norm_text(name)
        isin_s = isin_from_text(norm_text(isin)) or norm_text(isin)
        if name_s and isin_s:
            mapping[norm_key(name_s)] = isin_s

    if debug:
        print(f"Портфель: собрано записей {len(mapping)}")
    return mapping


def find_isin_by_name(text: str, name_to_isin: Dict[str, str]) -> Optional[str]:
    """
    Пытается найти ISIN инструмента по тексту, сравнивая нормализованные строки.
    Совпадение — если нормализованное имя из портфеля является подстрокой нормализованного text.
    """
    if not text:
        return None
    key = norm_key(text)
    for nk, isin in name_to_isin.items():
        if nk and nk in key:
            return isin
    return None

# ---------------------------- Сделки -----------------------------------------

def parse_trades_to_target(
    ws: Worksheet,
    name_to_isin: Dict[str, str],
    alloc_commission: bool = True,
    locale_comma: bool = True,
    debug: bool = False,
) -> List[TargetRow]:
    """
    Парсит раздел «СДЕЛКИ С ЦЕННЫМИ БУМАГАМИ» в список TargetRow (Buy/Sell).
    - пытается определить ISIN по заголовку выпуска (RU... или по портфелю);
    - комиссия = комиссия ТС + комиссия брокера; опционально распределяется по «итого по выпуску».
    """
    rows: List[TargetRow] = []

    title_row = find_section_title(ws, SECTION_TRADES_RE)
    if title_row is None:
        raise RuntimeError("Раздел «СДЕЛКИ С ЦЕННЫМИ БУМАГАМИ» не найден")

    header_row, idx_map = find_header_below(
        ws, title_row, HEADER_KEYS_TRADES,
        needed=["deal_date", "deal_type", "price", "qty", "amount", "amount_ccy"]
    )

    get = lambda name: idx_map.get(name, None)  # 0-based индекс
    current_issue_text: Optional[str] = None
    current_issue_isin: Optional[str] = None

    # Буфер сделок текущего выпуска для последующего распределения комиссий «итого по выпуску»
    batch_items: List[Tuple[int, float]] = []  # (индекс в rows, abs_amount)
    batch_total_fee: float = 0.0

    def flush_batch():
        nonlocal batch_items, batch_total_fee
        if alloc_commission and batch_items and batch_total_fee:
            total = sum(a for _, a in batch_items) or 1.0
            for idx_row, a in batch_items:
                share = batch_total_fee * (a / total)
                rows[idx_row].FeeTax = (rows[idx_row].FeeTax or 0.0) + share
        batch_items = []
        batch_total_fee = 0.0

    r = header_row + 1
    while r <= ws.max_row:
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        line_text = " | ".join(norm_text(v) for v in values)

        # Конец раздела (встретили заголовок другого раздела)
        if SECTION_CASH_RE.search(line_text) or SECTION_PORTFOLIO_RE.search(line_text) \
                or SECTION_CLEARANCE_RE.search(line_text):
            flush_batch()
            break

        # Строка «Итого по выпуску»
        if re.search(r"итого\s+по\s+выпуску", line_text, re.IGNORECASE):
            comm_ts = parse_decimal(values[get("comm_ts")] if get("comm_ts") is not None else None, locale_comma)
            comm_br = parse_decimal(values[get("comm_broker")] if get("comm_broker") is not None else None, locale_comma)
            batch_total_fee += (comm_ts or 0.0) + (comm_br or 0.0)
            flush_batch()
            r += 1
            continue

        # Шапка выпуска/эмитента — в колонке «Вид сделки» пусто
        deal_type_cell = values[get("deal_type")] if get("deal_type") is not None else None
        if norm_text(deal_type_cell) == "":
            head_text = " ".join(norm_text(values[c]) for c in range(len(values)) if norm_text(values[c]))
            if head_text:
                current_issue_text = head_text
                current_issue_isin = isin_from_text(head_text) or find_isin_by_name(head_text, name_to_isin)
            r += 1
            continue

        # Нормальная строка сделки
        deal_type = norm_text(values[get("deal_type")]).lower() if get("deal_type") is not None else ""
        deal_date = parse_date(values[get("deal_date")]) if get("deal_date") is not None else None
        if not deal_date or deal_type not in TRADE_TYPES_MAP:
            r += 1
            continue

        event = TRADE_TYPES_MAP[deal_type]
        price = parse_decimal(values[get("price")], locale_comma) if get("price") is not None else None
        qty = parse_decimal(values[get("qty")], locale_comma) if get("qty") is not None else None
        nkd = parse_decimal(values[get("nkd")], locale_comma) if get("nkd") is not None else None
        amount = parse_decimal(values[get("amount")], locale_comma) if get("amount") is not None else None
        ccy = normalize_currency(norm_text(values[get("amount_ccy")])) if get("amount_ccy") is not None else None

        fee_ts = parse_decimal(values[get("comm_ts")], locale_comma) if get("comm_ts") is not None else None
        fee_br = parse_decimal(values[get("comm_broker")], locale_comma) if get("comm_broker") is not None else None
        fee = (fee_ts or 0.0) + (fee_br or 0.0)
        fee = fee if fee != 0.0 else None

        place = norm_text(values[get("place")]) if get("place") is not None else None
        exchange = to_exchange(place)

        rows.append(TargetRow(
            Event=event,
            Date=deal_date,
            Symbol=current_issue_isin or None,
            Price=price,
            Quantity=qty,
            Currency=ccy,
            FeeTax=fee,
            Exchange=exchange,
            NKD=nkd or 0.0,
            FeeCurrency=None,
            DoNotAdjustCash=None,
            Note=current_issue_text or None
        ))

        # Если у сделки нет своей комиссии, возможно распределим общую «по выпуску»
        if alloc_commission and amount is not None and (fee is None or fee == 0.0):
            batch_items.append((len(rows) - 1, abs(amount)))

        r += 1

    return rows

# ---------------------------- Движение денежных средств ----------------------

def parse_cash_to_target(
    ws: Worksheet,
    name_to_isin: Dict[str, str],
    locale_comma: bool = True,
    map_coupon_as_price_one: bool = True,
    debug: bool = False,
) -> List[TargetRow]:
    """
    Парсит раздел «ДВИЖЕНИЕ ДЕНЕЖНЫХ СРЕДСТВ ...»:
    - игнорирует строки «Списано по сделке»;
    - «Ввод ДС» -> Cash_In (Symbol=валюта, Price=1, Quantity=сумма);
    - «Погашение купона» -> Dividend (пытаемся определить ISIN из комментария).
    При необходимости можно расширить обработку «Вывод ДС» -> Cash_Out.
    """
    rows: List[TargetRow] = []

    title_row = find_section_title(ws, SECTION_CASH_RE)
    if title_row is None:
        if debug:
            print("ДДС: раздел не найден.")
        return rows

    header_row, idx = find_header_below(ws, title_row, HEADER_KEYS_CASH, needed=["date", "type", "amount", "ccy"])
    col = {k: v for k, v in idx.items()}  # 1-based

    r = header_row + 1
    while r <= ws.max_row:
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(values):
            break

        op_type = norm_text(values[col["type"]]).lower() if "type" in col else ""
        date = parse_date(values[col["date"]]) if "date" in col else None
        amount = parse_decimal(values[col["amount"]], locale_comma) if "amount" in col else None
        ccy = normalize_currency(norm_text(values[col["ccy"]])) if "ccy" in col else None
        comment = norm_text(values[col["comment"]]) if "comment" in col else None

        if not date:
            r += 1
            continue

        # Игнорируем отражение денежных списаний по сделкам
        if "списано по сделке" in op_type or "списание по сделке" in op_type:
            r += 1
            continue

        # Ввод ДС -> Cash_In
        if "ввод дс" in op_type or "пополнение" in op_type:
            if amount:
                rows.append(TargetRow(
                    Event="Cash_In",
                    Date=date,
                    Symbol=ccy,
                    Price=1.0,
                    Quantity=abs(amount),
                    Currency=ccy,
                    FeeTax=0.0,
                    Exchange=None,
                    NKD=0.0,
                    FeeCurrency=None,
                    DoNotAdjustCash=None,
                    Note=comment
                ))
            r += 1
            continue

        # Возможная поддержка «Вывод ДС» -> Cash_Out (если встречается)
        if "вывод дс" in op_type or "снятие" in op_type:
            if amount:
                rows.append(TargetRow(
                    Event="Cash_Out",
                    Date=date,
                    Symbol=ccy,
                    Price=1.0,
                    Quantity=abs(amount),
                    Currency=ccy,
                    FeeTax=0.0,
                    Exchange=None,
                    NKD=0.0,
                    FeeCurrency=None,
                    DoNotAdjustCash=None,
                    Note=comment
                ))
            r += 1
            continue

        # Погашение купона -> Dividend
        if "погашение купона" in op_type or ("купон" in op_type) or ("купон" in (comment or "").lower()):
            symbol = isin_from_text(comment) or find_isin_by_name(comment or "", name_to_isin)
            # По спецификации Dividend: если известна только итоговая сумма,
            # корректно задать Price=1, Quantity=итог.
            price = 1.0 if map_coupon_as_price_one else None
            qty = abs(amount) if amount is not None else None
            rows.append(TargetRow(
                Event="Dividend",
                Date=date,
                Symbol=symbol,
                Price=price,
                Quantity=qty,
                Currency=ccy,
                FeeTax=0.0,  # налоги по купонам обычно в другой секции; можно расширить
                Exchange=None,
                NKD=0.0,
                FeeCurrency=None,
                DoNotAdjustCash=None,
                Note=comment
            ))
            r += 1
            continue

        # Иные типы операций сейчас пропускаем
        r += 1

    return rows